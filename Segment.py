import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from time import *


class Slice:
    pass


def input_data(inputFileRoute):
    #从文件读入数据到DataFrame
    data = pd.read_excel(inputFileRoute,header=0,sheet_name=0,usecols=[2,3,5,6,9,10,12,13,15,16])

    return data



def cal_resistance_data(data):
    #计算电阻数据
    for i in range(0,data.shape[1],2):
        I = (3.3 - data.iloc[:,i]) / 200000 #计算电流
        data.iloc[:,i] = (data.iloc[:,i] - data.iloc[:,i+1]) / I
        data.iloc[:,i+1] = data.iloc[:,i+1] / I
    
    #删除通道2
    data = data.drop(columns=['通道2'],axis=1)

    return data


def check_data(filename):
    global bottomLimit
    global topLimit

    checkData = pd.read_excel(filename,header=0,sheet_name=0,usecols=[2,3,5,6,9,10,12,13,15,16])

    checkData = cal_resistance_data(checkData)

    dataSlicePoint = segment_resistance_data(checkData)
    bottomLimit = checkData.iloc[dataSlicePoint[0],:].values.tolist()
    topLimit = checkData.iloc[dataSlicePoint[1],:].values.tolist()
    print(bottomLimit)
    print(topLimit)




def cal_data(data):
    global bottomLimit
    global topLimit

    #计算边界值
    check_data(r"./第二次手套数据/手套1/第5组/处理后的数据集/P_手套1-5次-校正.xlsx")

    #计算电阻数据
    for i in range(0,data.shape[1],2):
        I = (3.3 - data.iloc[:,i]) / 200000 #计算电流
        data.iloc[:,i] = (data.iloc[:,i] - data.iloc[:,i+1]) / I
        data.iloc[:,i+1] = data.iloc[:,i+1] / I
    
    #删除通道2
    data = data.drop(columns=['通道2'],axis=1)

    #将数据规范到最小值和最大值之间并进行归一化
    for i in range(data.shape[1]):
        data.iloc[:,i] = np.clip(data.iloc[:,i],bottomLimit[i],topLimit[i])
        data.iloc[:,i] = (data.iloc[:,i] - bottomLimit[i]) / (topLimit[i] - bottomLimit[i])

    return data


    




def dynamic_segment_data(data):
    #定义一些参数
    FrameLength = 10
    StepLength = 4
    Threshold1 = 0.010
    Threshold2 = 0.003

    #计算波动值序列
    fluctuation = np.zeros(data.shape[0]-FrameLength)
    for i in range(data.shape[0]-FrameLength):
        fluctuation[i] = np.sum(np.var(data.iloc[i:i+FrameLength+1,:]))
    
    # 观察波动值变化情况
    plt.plot(fluctuation)
    plt.xticks(np.arange(0,len(fluctuation),40))
    plt.axhline(y=0.003,c='red',ls='--',lw=1)
    plt.axhline(y=0.010,c='red',ls='--',lw=1)
    plt.show()

    findStart = False
    slices = []
    sliceTmp = Slice()
    validIndex = []
    #遍历波动序列进行分割
    for i in  range(0,fluctuation.shape[0],StepLength):
        if not findStart and fluctuation[i] >= Threshold1:
            sliceTmp.start = i
            findStart = True
        elif findStart and (fluctuation[i] < Threshold2):
            sliceTmp.end = i
            findStart =  False
            slices.append(sliceTmp)
            validIndex = np.r_[validIndex,sliceTmp.start:sliceTmp.end]
            sliceTmp = Slice()
    
    #根据有效序列对原始数据进行切片
    data.plot(xticks= np.arange(0,data.shape[0],40),grid=True)
    for val in slices:
        if val.start > 1000:
            break
        plt.axvline(x=val.start,c='red',ls='--',lw=2)
        plt.axvline(x=val.end,c='blue',ls='--',lw=2)
    plt.show()

    #将分割出的第一个手势序列存入models.xlsx中作为模板
    # book = load_workbook(r"./手势分割数据/models.xlsx") #加载工作簿
    # writer = pd.ExcelWriter(r"./手势分割数据/models.xlsx") #打开一个excel文件
    # writer.book = book
    # data.iloc[slices[2].start:slices[2].end+1,:].to_excel(writer,startrow=0,index=False,sheet_name='Sheet4') #将分割出来的第一段序列写入到指定位置
    # writer.save() #保存

    return slices



def static_segment_data(data):
    #定义一些参数
    FrameLength = 15
    StepLength = 10
    Threshold1 = 0.015
    Threshold2 = 0.015

    #计算波动值序列
    fluctuation = np.zeros(data.shape[0]-FrameLength)
    for i in range(data.shape[0]-FrameLength):
        fluctuation[i] = np.sum(np.var(data.iloc[i:i+FrameLength+1,:]))
    
    # print(max(fluctuation[800:960]))

    # 观察波动值变化情况
    plt.plot(fluctuation)
    plt.xticks(np.arange(0,len(fluctuation),40))
    plt.axhline(y=0.015,c='red',ls='--',lw=1)
    plt.axhline(y=0.015,c='red',ls='--',lw=1)
    plt.show()

    findStart = False
    slices = []
    sliceTmp = Slice()
    validIndex = []
    #遍历波动序列进行分割
    for i in  range(0,fluctuation.shape[0],StepLength):
        if not findStart and fluctuation[i] < Threshold1:
            sliceTmp.start = i
            findStart = True
        elif findStart and (fluctuation[i] > Threshold2):
            sliceTmp.end = i
            findStart =  False
            slices.append(sliceTmp)
            validIndex = np.r_[validIndex,sliceTmp.start:sliceTmp.end]
            sliceTmp = Slice()

    #根据切片数组选取每个切片中间的数据作为代表，生成分割数据
    dataSlicePoint = []
    for i in range(len(slices)):
        dataSlicePoint.append((slices[i].start+slices[i].end) // 2)
    

    # 根据有效序列对原始数据进行切片
    data.plot(xticks= np.arange(0,data.shape[0],200),grid=True)
    for val in slices:
        plt.axvline(x=val.start,c='red',ls='--',lw=2)
        plt.axvline(x=val.end,c='blue',ls='--',lw=2)
    plt.show()


    # #将分割出的第一个手势序列存入models.xlsx中作为模板
    # book = load_workbook(r"./手套数据/models.xlsx") #加载工作簿
    # writer = pd.ExcelWriter(r"./手套数据/models.xlsx") #打开一个excel文件
    # writer.book = book

    # # averModelData = pd.DataFrame(np.zeros((3,9)),columns=['通道1','通道4','通道5','通道8','通道9','通道11','通道12','通道14','通道15'])
    # # for i in range(len(dataSlicePoint)):
    # #     averModelData = averModelData + data.iloc[[slices[i].start,dataSlicePoint[i],slices[i].end],:].reset_index(drop=True)
    # # print(averModelData)
    # # averModelData /= len(dataSlicePoint)

    # data.iloc[[slices[1].start,dataSlicePoint[1],slices[1].end],:].to_excel(writer,startrow=0,index=False,sheet_name='Sheet1') #将分割出来的第一段序列写入到指定位置
    # writer.save() #保存

    return dataSlicePoint


def segment_resistance_data(data):
    #定义一些参数
    FrameLength = 15
    StepLength = 10
    Threshold1 = 0.3e9
    Threshold2 = 0.5e9

    #计算波动值序列
    fluctuation = np.zeros(data.shape[0]-FrameLength)
    for i in range(data.shape[0]-FrameLength):
        fluctuation[i] = np.sum(np.var(data.iloc[i:i+FrameLength+1,:]))
    
    # print(max(fluctuation[800:960]))

    # 观察波动值变化情况
    plt.plot(fluctuation)
    plt.xticks(np.arange(0,len(fluctuation),40))
    plt.axhline(y=0.3e9,c='red',ls='--',lw=1)
    plt.axhline(y=0.5e9,c='red',ls='--',lw=1)
    plt.show()

    findStart = False
    slices = []
    sliceTmp = Slice()
    #遍历波动序列进行分割
    for i in  range(0,fluctuation.shape[0],StepLength):
        if not findStart and fluctuation[i] < Threshold1:
            sliceTmp.start = i
            findStart = True
        elif findStart and (fluctuation[i] > Threshold2):
            sliceTmp.end = i
            findStart =  False
            slices.append(sliceTmp)
            sliceTmp = Slice()

    #根据切片数组选取每个切片中间的数据作为代表，生成分割数据
    dataSlicePoint = []
    for i in range(len(slices)):
        dataSlicePoint.append((slices[i].start+slices[i].end) // 2)
    

    # 根据有效序列对原始数据进行切片
    data.plot(xticks= np.arange(0,data.shape[0],200),grid=True)
    for val in slices:
        plt.axvline(x=val.start,c='red',ls='--',lw=2)
        plt.axvline(x=val.end,c='blue',ls='--',lw=2)
    plt.show()

    return dataSlicePoint



global topLimit
topLimit = [395896.790434,72071.153410,101824.155903,143580.095946,
277245.790612,114187.643021,189244.851259,270555.498299,209646.501082]

global bottomLimit
bottomLimit = [260628.875111,48835.917378,84480.581789,79084.237396,
247399.489470,79435.589234,151711.523386,163810.162670,189602.549052]


if __name__ == '__main__':
    begin_time = time()
    #输入数据
    inputFileRoute = r"./第二次手套数据/手套1/第5组/处理后的数据集/P_手套1-5次-组合-1-2-3-4-0-握拳-0-握拳-0-握拳-0-4-3-2-1.xlsx"
    data = input_data(inputFileRoute)

    # 处理数据集
    data = cal_data(data)


    # 这两行代码使得 pyplot 画出的图形中可以显示中文
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    # #观察原始数据分布情况
    # data.plot(xticks=np.arange(0,data.shape[0],40),grid=True)
    # plt.show()
    # print(data.describe())
    # print(np.argmax(data.values,axis=0))

    # pdata = data.cumsum().apply(lambda x: x / (int(x.name)+1),axis=1)
    # pdata = (pdata - pdata.min())/(pdata.max() - pdata.min())
    # pdata.plot(xticks= np.arange(0,pdata.shape[0],40),grid=True)
    # plt.show()


    static_segment_data(data)
    end_time = time()
    print("该程序运行时间为{:.5f}秒".format(end_time-begin_time))




